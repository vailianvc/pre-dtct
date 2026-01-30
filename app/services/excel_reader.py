import os
from openpyxl import load_workbook
from app import db
from app.models import GlossaryCache

def load_glossary(file_path, glossary_type):
    """
    Read a glossary Excel file and return data as list of dicts

    Args:
        file_path: Path to Excel file
        glossary_type: Type of glossary (academicsession, programme, etc.)

    Returns:
        List of dicts with 'code' and 'description' keys
        For academicsession type, also includes 'commencement_week_1' and 'commencement_week_2'
    """
    try:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            # Activity glossary has columns swapped: Activity Name (col 0), Activity Code (col 1)
            if glossary_type == 'activity':
                code = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                description = str(row[0]).strip() if row[0] else ''
                item = {'code': code, 'description': description}
            elif glossary_type == 'academicsession':
                # Academic session has additional columns: Commencement Week 1 (col 2), Week 2 (col 3)
                code = str(row[0]).strip() if row[0] else ''
                description = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                week1 = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                week2 = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                item = {
                    'code': code,
                    'description': description,
                    'commencement_week_1': week1,
                    'commencement_week_2': week2
                }
            else:
                code = str(row[0]).strip() if row[0] else ''
                description = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                item = {'code': code, 'description': description}

            if code:
                data.append(item)

        wb.close()
        return data
    except Exception as e:
        print(f"Error loading glossary {file_path}: {e}")
        return []

def load_all_glossaries(app):
    """
    Load all glossary files into database cache

    Args:
        app: Flask application instance
    """
    glossary_dir = app.config['GLOSSARY_DIR']
    glossary_files = app.config['GLOSSARY_FILES']

    # Check if cache is already populated
    existing_count = GlossaryCache.query.count()
    if existing_count > 0:
        print(f"Glossary cache already populated with {existing_count} entries")
        return

    print("Loading glossary files into database cache...")

    for glossary_type, filename in glossary_files.items():
        file_path = os.path.join(glossary_dir, filename)

        if not os.path.exists(file_path):
            print(f"Warning: Glossary file not found: {file_path}")
            continue

        data = load_glossary(file_path, glossary_type)

        # Insert into database (skip duplicates)
        inserted_count = 0
        for item in data:
            # Check if entry already exists
            existing = GlossaryCache.query.filter_by(
                glossary_type=glossary_type,
                code=item['code']
            ).first()

            if not existing:
                entry = GlossaryCache(
                    glossary_type=glossary_type,
                    code=item['code'],
                    description=item['description'],
                    commencement_week_1=item.get('commencement_week_1'),
                    commencement_week_2=item.get('commencement_week_2')
                )
                db.session.add(entry)
                inserted_count += 1

        print(f"Loaded {inserted_count} entries for {glossary_type} ({len(data)} total in file)")

    try:
        db.session.commit()
        total_count = GlossaryCache.query.count()
        print(f"Glossary cache populated successfully with {total_count} total entries")
    except Exception as e:
        db.session.rollback()
        print(f"Error populating glossary cache: {e}")

def get_glossary_data(glossary_type):
    """
    Get glossary data from database cache

    Args:
        glossary_type: Type of glossary to retrieve

    Returns:
        List of dicts with 'code' and 'description' keys
    """
    entries = GlossaryCache.query.filter_by(glossary_type=glossary_type).all()
    return [entry.to_dict() for entry in entries]
